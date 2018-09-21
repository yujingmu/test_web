#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/7/26 22:25
# @Author  : yimi
# @File    : get_excel_data.py
# 先读初始化数据-->处理数据-->对数据进行替换-->更新初始化数据
from openpyxl import load_workbook
from task_0810_database.common.read_config_data import ReadConfigDate
from task_0810_database.config_file import read_path

class GetDataFromExcel:
    def __init__(self, path_file, sheet_name, mode, case_id_list):
        self.path_file = path_file
        self.sheet_name = sheet_name
        self.mode = mode
        self.case_id_list = case_id_list

    def get_data_from_excel(self):
        print('===============%s===============',self.path_file)
        wb = load_workbook(self.path_file)
        sheet = wb[self.sheet_name]
        print('===================%s===========',self.sheet_name)
        row = sheet.max_row
        no_reg_tel = self.get_init_data()
        no_reg_tel_new = self.get_init_data() + 1
        row_list = []
        if self.mode == '0':
            for i in eval(self.case_id_list): #[1,3]
                m = i + 1
                row_dict = {}
                row_dict['case_id'] = sheet.cell(m, 1).value
                row_dict['module'] = sheet.cell(m, 2).value
                row_dict['title'] = sheet.cell(m, 3).value
                row_dict['method'] = sheet.cell(m, 4).value
                row_dict['url'] = sheet.cell(m, 5).value

                if str(sheet.cell(m, 6).value).find('${no_reg_tel}') != -1:
                    new_param = sheet.cell(m, 6).value.replace('${no_reg_tel}', str(no_reg_tel + 1))
                    row_dict['param'] = new_param
                else:
                    row_dict['param'] = sheet.cell(m, 6).value
                row_dict['ExpectedResult'] = sheet.cell(m, 7).value
                row_dict['CheckSQL'] = sheet.cell(m, 8).value
                row_list.append(row_dict)
        else:
            for i in range(2, row + 1):
                row_dict = {}
                row_dict['case_id'] = sheet.cell(i, 1).value
                row_dict['module'] = sheet.cell(i, 2).value
                row_dict['title'] = sheet.cell(i, 3).value
                row_dict['method'] = sheet.cell(i, 4).value
                row_dict['url'] = sheet.cell(i, 5).value
                if str(sheet.cell(i, 6).value).find('${no_reg_tel}') != -1:
                    new_param = sheet.cell(i, 6).value.replace('${no_reg_tel}', str(no_reg_tel + 1))
                    row_dict['param'] = new_param
                else:
                    row_dict['param'] = sheet.cell(i, 6).value
                row_dict['ExpectedResult'] = sheet.cell(i, 7).value
                row_dict['CheckSQL'] = sheet.cell(i, 8).value
                row_list.append(row_dict)
        self.update_init_date(no_reg_tel_new)
        return row_list

    def get_init_data(self):
        wb=load_workbook(self.path_file)
        sheet=wb['init']
        no_reg_tel=sheet.cell(1, 2).value
        return no_reg_tel

    def update_init_date(self, no_reg_tel):
        wb = load_workbook(self.path_file)
        sheet = wb['init']
        sheet.cell(1, 2).value = no_reg_tel
        wb.save(self.path_file)

    def write_data(self, row, col, actual_result):
        wb = load_workbook(self.path_file)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = actual_result
        wb.save(self.path_file)

if __name__ == "__main__":
    mode = ReadConfigDate(read_path.config_path, 'FLAG', 'mode').read_config()
    case_id_list = ReadConfigDate(read_path.config_path, 'FLAG', 'case_id').read_config()
    test_date = GetDataFromExcel(read_path.test_data_path, 'test_data', mode, case_id_list).get_data_from_excel()
    print(test_date)


